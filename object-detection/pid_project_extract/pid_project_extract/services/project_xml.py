from interval import Interval
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

from pid_project_extract.services.utils import XmlProjectItem


class PIDProjectXml(Workbook):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.project_list = None
        self.xml_projects = []
        self.current_project_info_list = None
        self.current_proofreader_info_list = None

    def construct(self, projects_info):
        """
        解析出来如果正确，第一个一定是项目名称；
        在项目名称下，是装置名称，获取规格是，计算名称的【项目名称】的右下[2]与左下[3]坐标的x轴平均值，
            然后在其 y 轴上增加 80（经验值），判断包含了这个坐标的数据项即为【装置名称】
        在装置名称下，是逻辑图名称，算法同上。
        """

        if projects_info is None:
            print("projects info is None")
            return

        self.project_list = projects_info
        for item in self.project_list.items():
            print("Construct xml data: ", item[0])
            self.current_project_info_list, self.current_proofreader_info_list = item[1]
            project_info_list = self.current_project_info_list

            project_no_item = get_project_no_item(project_info_list)
            if project_no_item is None:
                print("project number item no exists")
                continue

            if len(project_info_list) >= 5:
                project_name = project_info_list[0].value
                project_pos = project_info_list[0].position

                project_no, project_no_pos = self.get_project_no(
                    project_no_item.position,
                    project_info_list)
                device_name, device_pos = get_device_name(project_pos,
                                                          project_info_list)

                design_phase = "详细工程设计"
                contract_no = ""
                pic_no, pic_pos = self.get_pic_no(project_no_pos, project_info_list)
                logic_name, logic_pos = self.get_logic_name(device_pos,
                                                            project_info_list)
                desc_val, desc_pos = self.get_description()
                description = desc_val if desc_val is not None else ""
                designer = ""
                checker = ""
                reviewer = ""
                approve = ""
                date, data_pos = self.get_date()
                self.xml_projects.append(
                    XmlProjectItem(project_name, project_no, device_name, design_phase,
                                   pic_no,
                                   logic_name, date, contract_no, description, designer,
                                   checker, reviewer, approve))

    def save(self, filename):
        sheet = self.active

        # 写入表头
        sheet.append(
            ['项目名称', '项目号', '装置名称', '设计阶段', '合同号', '图号',
             '逻辑图名称',
             '说明', '设计', '校核', '审核', '审定', '日期'])

        # 遍历对象数组并将数据写入Excel表格
        for item in self.xml_projects:
            print("write data to sheet: ", item.project_name)
            sheet.append([item.project_name, item.project_no, item.device_name,
                          item.design_phase, item.contract_no, item.pic_no,
                          item.logic_name, item.description, item.designer,
                          item.checker, item.reviewer, item.approve, item.date])

        # 设置单元格边框
        thin_border = Border(left=Side(style="thin"),
                             right=Side(style="thin"),
                             top=Side(style="thin"),
                             bottom=Side(style="thin"))

        for row in sheet.iter_rows(min_row=1, max_row=len(self.xml_projects) + 1,
                                   min_col=1,
                                   max_col=13):
            for cell in row:
                cell.border = thin_border

        # 设置表头的加粗样式
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # 设置表头单元格的加粗样式
        for cell in sheet[1]:
            cell.font = bold_font
            cell.alignment = center_alignment

        # 自适应设置列宽度
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # 获取列的字母标识符
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.5  # 调整为合适的宽度，这里的 1.5 是一个经验值
            sheet.column_dimensions[column_letter].width = adjusted_width

        print("save xml file: ", filename)
        super().save(filename)

    def get_logic_name(self, device_name_pos, project_info_list):
        if device_name_pos is None:
            return "", None

        x_interval = Interval(device_name_pos[2][0], device_name_pos[3][0])
        y_interval = Interval(device_name_pos[2][1], device_name_pos[2][1] + 120)

        for item in project_info_list:
            if item.position != device_name_pos and is_inside_rect(item.position,
                                                                   x_interval,
                                                                   y_interval):
                return item.value, item.position

        return "", None

    def get_project_no(self, project_no_pos, project_info_list):
        if project_no_pos is None:
            return "", None

        x_interval = Interval(project_no_pos[2][0], project_no_pos[2][0] + 50)
        y_interval = Interval(project_no_pos[1][1], project_no_pos[2][1])

        for item in project_info_list:
            if item.position != project_no_pos and is_inside_rect(item.position,
                                                                  x_interval,
                                                                  y_interval):
                return item.value, item.position

        return "", None

    def get_pic_no(self, project_no_pos, project_info_list):
        if project_no_pos is None:
            return "", None

        x_interval = Interval(project_no_pos[2][0], project_no_pos[3][0])
        y_interval = Interval(project_no_pos[2][1], project_no_pos[2][1] + 30)

        for item in project_info_list:
            if item.position != project_no_pos and is_inside_rect(item.position,
                                                                  x_interval,
                                                                  y_interval):
                return item.value, item.position

        return "", None

    def get_description(self):
        x_interval = None
        y_interval = None
        for item in self.current_proofreader_info_list:
            if item.value == "说明":
                # 往上找 50 个像素
                x_interval = Interval(item.position[0][0], item.position[1][0])
                y_interval = Interval(item.position[0][1], item.position[0][1] - 50)
                break
        if x_interval is not None:
            for item in self.current_proofreader_info_list:
                if item.value != "说明" and is_inside_rect(item.position,
                                                           x_interval, y_interval):
                    return item.value, item.position
        return "", None

    def get_date(self):
        x_interval = None
        y_interval = None
        for item in self.current_proofreader_info_list:
            if item.value == "日期":
                # 往上找 50 个像素
                x_interval = Interval(item.position[0][0], item.position[1][0])
                y_interval = Interval(item.position[0][1], item.position[0][1] - 50)
                break
        if x_interval is not None:
            for item in self.current_proofreader_info_list:
                if item.value != "日期" and is_inside_rect(item.position,
                                                           x_interval, y_interval):
                    return item.value, item.position
        return "", None


def is_inside_rect(rect, target_x_interval: Interval,
                   target_y_interval: Interval):
    x1, y1 = rect[0]
    x2, y2 = rect[2]
    x_interval = Interval(x1, x2)
    y_interval = Interval(y1, y2)
    return target_x_interval.overlaps(x_interval) and target_y_interval.overlaps(
        y_interval)


def get_device_name(project_name_pos, project_info_list):
    if project_name_pos is None:
        return "", None

    x_interval = Interval(project_name_pos[2][0], project_name_pos[3][0])
    y_interval = Interval(project_name_pos[2][1], project_name_pos[2][1] + 80)

    for item in project_info_list:
        if item.position != project_name_pos and is_inside_rect(item.position,
                                                                x_interval,
                                                                y_interval):
            return item.value, item.position
    return "", None


def get_project_no_item(project_info_list):
    for item in project_info_list:
        if item.value == "项目号":
            return item
